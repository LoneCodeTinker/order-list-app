from fastapi import FastAPI, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
import os
from dotenv import load_dotenv
import json
import openpyxl
from datetime import datetime
import glob

load_dotenv()

app = FastAPI()

# Allow CORS for local frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "*"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# In-memory inventory (barcode -> {name, price})
inventory = {}

UPLOAD_DIR = os.path.join(os.getcwd(), "uploaded_inventory")
ORDER_SAVE_DIR = os.getenv("ORDER_SAVE_DIR", os.path.join(os.getcwd(), "orders"))
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(ORDER_SAVE_DIR, exist_ok=True)

def get_latest_inventory_file():
    files = glob.glob(os.path.join(UPLOAD_DIR, '*.xls*'))
    if not files:
        return None
    return max(files, key=os.path.getctime)

@app.post("/upload-inventory")
def upload_inventory(file: UploadFile = File(...)):
    # Only accept .xls or .xlsx
    if not (file.filename.endswith('.xls') or file.filename.endswith('.xlsx')):
        return JSONResponse(status_code=400, content={"error": "Only .xls and .xlsx files are allowed."})
    # Save uploaded file with timestamp
    now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    ext = os.path.splitext(file.filename)[1]
    new_filename = f"{now}{ext}"
    file_path = os.path.join(UPLOAD_DIR, new_filename)
    with open(file_path, "wb") as f:
        f.write(file.file.read())
    # Load inventory into memory
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    inventory.clear()
    for row in ws.iter_rows(min_row=2, values_only=True):
        barcode, name, price = row[0], row[1], row[2] if len(row) > 2 else (None)
        if barcode:
            inventory[str(barcode)] = {"name": name, "price": price if price is not None else 0.0}
    return {"count": len(inventory), "saved_as": file_path}

@app.get("/latest-inventory")
def latest_inventory():
    latest_file = get_latest_inventory_file()
    if not latest_file:
        return JSONResponse(status_code=404, content={"error": "No inventory file found."})
    # Load inventory into memory
    wb = openpyxl.load_workbook(latest_file)
    ws = wb.active
    inventory.clear()
    for row in ws.iter_rows(min_row=2, values_only=True):
        barcode, name, price = row[0], row[1], row[2] if len(row) > 2 else (None)
        if barcode:
            inventory[str(barcode)] = {"name": name, "price": price if price is not None else 0.0}
    return {"count": len(inventory), "filename": os.path.basename(latest_file)}

@app.get("/item/{barcode}")
def get_item(barcode: str):
    # Always use latest inventory file
    latest_file = get_latest_inventory_file()
    if not latest_file:
        return JSONResponse(status_code=404, content={"error": "No inventory file found."})
    wb = openpyxl.load_workbook(latest_file)
    ws = wb.active
    temp_inventory = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        barcode_val, name, price = row[0], row[1], row[2] if len(row) > 2 else (None)
        if barcode_val:
            temp_inventory[str(barcode_val)] = {"name": name, "price": price if price is not None else 0.0}
    item = temp_inventory.get(barcode)
    if item:
        return item
    return JSONResponse(status_code=404, content={"error": "Item not found"})

@app.post("/save-order")
def save_order(
    customer_name: str = Form(...),
    customer_phone: str = Form(...),
    username: str = Form(...),
    items: str = Form(...),
    created_by: str = Form(...),
):
    from openpyxl.styles import Alignment
    items_list = json.loads(items)
    now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    # Clean filename for Windows
    def clean_filename(s):
        return "".join(c for c in s if c.isalnum() or c in (' ', '-', '_')).rstrip()
    safe_customer = clean_filename(customer_name)
    safe_username = clean_filename(username)
    safe_created_by = clean_filename(created_by)
    filename = f"{now}_{safe_customer}_{safe_username}-created by {safe_created_by}.xlsx"
    filepath = os.path.join(ORDER_SAVE_DIR, filename)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Order"
    ws.append(["Customer Name", customer_name])
    ws.append(["Phone Number", customer_phone])
    ws.append(["Username", username])
    ws.append(["Created By", created_by])
    ws.append(["Date", now])
    ws.append([])
    ws.append(["Barcode", "Name", "Quantity", "Price", "Total", "VAT (15%)"])
    order_total = 0.0
    order_vat = 0.0
    for item in items_list:
        price = float(item.get("price", 0.0))
        quantity = int(item["quantity"])
        total = price * quantity
        vat = total * 0.15
        order_total += total
        order_vat += vat
        ws.append([
            item["barcode"],
            item["name"],
            quantity,
            price,
            total,
            vat
        ])
    ws.append([])
    ws.append(["Order Total", order_total])
    ws.append(["Order VAT (15%)", order_vat])
    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
        for cell in col:
            cell.alignment = Alignment(vertical="center", horizontal="center")
    wb.save(filepath)
    return {"filename": filename, "path": filepath}
